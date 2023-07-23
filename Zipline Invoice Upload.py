from tkinter import messagebox
from PyQt5.QtCore import QTimer
from PyQt5.QtWidgets import QWidget, QLineEdit, QLabel, QPushButton, QFormLayout, QHBoxLayout
import subprocess
from tkinter import ttk
import pyperclip
import webbrowser
import pygetwindow as gw
import tkinter as tk
from PyQt5.QtWidgets import QApplication
from PyQt5.QtCore import Qt
import re
import pdfplumber
import pandas as pd
import win32com.client
import keyboard
import os
import csv
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import time


def show_message_box(title, message):
    # We need to create a root window and immediately hide it
    # because a tkinter messagebox is associated with a parent window
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.update()
    messagebox.showinfo(title, message)
    root.destroy()

# Step 1: Rename pdf files to correct invoice number

show_message_box("Attention", "Please select all emails with Zipline invoices that you want to save, then press ` to continue...")

# Wait for the Enter key to be pressed globally
keyboard.wait('`')

source_dir = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Invoices"

# Iterate through all the files in the source directory
for file_name in os.listdir(source_dir):
    file_path = os.path.join(source_dir, file_name)

    # Check if the file is a PDF
    if file_path.lower().endswith(".pdf"):
        # Read the PDF file
        with pdfplumber.open(file_path) as pdf:
            # Extract the text from all pages
            text = "\n".join([page.extract_text() for page in pdf.pages])

        # Find the second occurrence of the phrase "Powered By McLeod Software"
        phrase = "Powered By McLeod Software"
        second_occurrence_index = text.find(phrase, text.find(phrase) + 1)

        # Use a regular expression to find the invoice number after the second occurrence of the phrase
        invoice_number_match = re.search(r'(\d{7})', text[second_occurrence_index + len(phrase):])
        if invoice_number_match:
            # Extract the invoice number and remove the leading 0
            invoice_number = invoice_number_match.group(1).lstrip('0')

            # Rename the PDF file with the updated invoice number
            new_file_path = os.path.join(source_dir, f"{invoice_number}.pdf")
            os.rename(file_path, new_file_path)

            print(f"File {file_name} renamed to {invoice_number}.pdf")
        else:
            print(f"Failed to find invoice number in {file_name}")


# Step 2: Get attachment with invoice report from Outlook

# Taking this out just for right now, step two can be reinserted after testing is complete

# Step 3
try:
    # Load the workbook and select the sheet
    wb = load_workbook(filename=r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")
    ws = wb["Data"]
except Exception as e:
    print(f"An error occurred: {e}")

# Delete the first two rows
ws.delete_rows(1, 2)
# Add 'Charge Type' to cell O1 and 'Charge Amount' to cell P1
ws['O1'] = 'Charge Type'
ws['P1'] = 'Charge Amount'

# Iterate over each row in the worksheet
for row in ws.iter_rows(min_row=2, max_col=15, max_row=ws.max_row):
    zipline_order = row[0].value  # Column A

    # If Column A is not empty, we have a new invoice
    if zipline_order is not None and zipline_order.strip() != "":
        freight_charge_label = 'Freight Charge'
        freight_charge_value = row[12].value  # Column M

        # Place 'Freight Charge' in column O and the value from column M in column P
        ws.cell(row=row[0].row, column=15, value=freight_charge_label)  # Column O
        ws.cell(row=row[0].row, column=16, value=freight_charge_value)  # Column P

# Delete columns M, N, Q, S, T, and U
# Note: Delete from right to left to keep indices correct
for col in ["U", "T", "S", "Q", "N", "M"]:
    ws.delete_cols(ws[col][0].column)

# Save the workbook
wb.save(filename=r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")


# Step 4: Fill invoice data down for all charges:

path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx"

# Load your workbook and select the 'Data' sheet
wb = load_workbook(filename = path)
ws = wb['Data']

# Columns to forward fill. Adjust according to your needs
columns_to_fill = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 14]  # 0-indexed

# Forward fill data for specified columns
current_values = [None] * len(columns_to_fill)
for row in ws.iter_rows(min_row=2):  # start from 2 to exclude header
    for i, column in enumerate(columns_to_fill):
        if row[column].value:  # if cell is not empty
            current_values[i] = row[column].value
        else:  # if cell is empty
            row[column].value = current_values[i]

# Note: Delete from right to left to keep indices correct
for col in ["H", "D", "C"]:
    ws.delete_cols(ws[col][0].column)

for row in ws.iter_rows(min_row=2):  # assuming first row is header
    cell = row[0]  # assuming the column you want to change is 'A'
    try:
        cell.value = int(cell.value)  # convert to integer
    except ValueError:
        try:
            cell.value = float(cell.value)  # convert to float
        except ValueError:
            pass  # not a number, do nothing

# Save the workbook
wb.save(path)

# Step 5: Create form to get Order Number and PO Number from PDF, and then replace the PO Number and Order number in the excel spreadsheet


def create_form(invoice_number, order_numbers, po_numbers, submit_callback):
    class MainWindow(QWidget):
        def __init__(self):
            super().__init__()
            # Set window title and minimum size
            self.setWindowTitle('Invoice Form')
            self.setMinimumSize(300, 200)
            self.move(300, 400)  # Position the window

            # Set the window to always stay on top
            self.setWindowFlags(Qt.WindowStaysOnTopHint)

            self.clipboard = QApplication.clipboard()
            # Create layout
            layout = QFormLayout()
            layout.setSpacing(20)
            layout.setContentsMargins(10,10,10,10)
            # Create invoice number label and copy button, and add them to the layout
            invoice_label = QLabel(f'Invoice Number: {invoice_number}')

            copy_button = QPushButton('Copy')
            copy_button.setMinimumHeight(30)  # Change 30 to whatever value works best
            copy_button.clicked.connect(lambda invoice_number=invoice_number: self.copy_to_clipboard(invoice_number))

            invoice_layout = QHBoxLayout()
            invoice_layout.addWidget(invoice_label)
            invoice_layout.addWidget(copy_button)
            layout.addRow(invoice_layout)

            # Store the QLineEdit widgets for order numbers and PO numbers
            self.order_number_widgets = []
            self.po_number_widgets = []

            # Create widgets and add them to layout
            for i, order_number in enumerate(order_numbers):
                order_number_widget = QLineEdit(order_number)
                layout.addRow(QLabel(f'Order Number {i + 1}:'), order_number_widget)
                self.order_number_widgets.append(order_number_widget)

            for i, po_number in enumerate(po_numbers):
                po_number_widget = QLineEdit(po_number)
                layout.addRow(QLabel(f'PO Number {i+1}:'), po_number_widget)
                self.po_number_widgets.append(po_number_widget)

            # Create a push button
            button = QPushButton('Submit')
            button.setDefault(True)
            layout.addRow(button)

            # Connect button click to handler
            button.clicked.connect(self.handle_button_click)

            # Set layout to the window
            self.setLayout(layout)

            # Adjust size to fit contents
            self.adjustSize()

            # Set stylesheet for the window
            self.setStyleSheet("""
                QWidget {
                    font-size: 15px;
                }
                QLabel {
                    color: #5a5a5a;
                }
                QLineEdit {
                    border: 1px solid #b5b5b5;
                    border-radius: 5px;
                    padding: 5px;
                }
                QPushButton {
                    border: 1px solid #5a5a5a;
                    border-radius: 5px;
                    padding: 5px;
                    background-color: #5a5a5a;
                    color: #ffffff;
                }
                QPushButton:hover {
                    border: 1px solid #7f7f7f;
                    background-color: #7f7f7f;
                }
                QPushButton:pressed {
                    background-color: #3a3a3a;
                }
                """)

        def handle_button_click(self):
            # Get the values from the QLineEdit widgets
            order_numbers = [widget.text() for widget in self.order_number_widgets]
            po_numbers = [widget.text() for widget in self.po_number_widgets]

            # Call the submit callback with the values
            submit_callback(invoice_number, order_numbers, po_numbers)

        def copy_to_clipboard(self, text):

            pyperclip.copy(invoice_number)

    return MainWindow

class FormManager:
    def __init__(self, data):
        self.dataframe = data  # Store the original DataFrame
        self.data = iter(data)  # Change this line
        self.current_window = None
        self.current_process = None  # Store reference to the current subprocess

        # Load the workbook once and keep it in memory
        self.workbook = load_workbook(filename=r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")
        self.worksheet = self.workbook.active

    def start(self):
        self.show_next_form()

    def show_next_form(self):
        try:
            row = next(self.data)
        except StopIteration:
            # No more data
            if self.current_window is not None:
                self.current_window.close()
            return  # Return immediately without trying to close NitroPDF

        # Extract and split the Customer Order Numbers and PO & Confirmation Numbers
        order_numbers = [num.strip() for num in str(row['Cust Order #']).split(',')]
        po_numbers = [num.strip() for num in str(row['PO and Confirmation Number']).split(',')]

        # Create a form with the extracted numbers
        window_class = create_form(row['Zipline Order #'], order_numbers, po_numbers, self.handle_form_submit)

        # Close the current window if it exists
        if self.current_window is not None:
            self.current_window.close()

        # Show the new window
        self.current_window = window_class()
        self.current_window.show()

        # Open the corresponding PDF file in NitroPDF
        self.current_process = self.open_pdf(row['Zipline Order #'])

    def handle_form_submit(self, invoice_number, order_numbers, po_numbers):
        # Join the values with commas
        order_numbers_str = ', '.join(order_numbers)
        po_numbers_str = ', '.join(po_numbers)

        for row in self.worksheet.iter_rows(min_row=2):  # Assuming first row is header
            if row[0].value == invoice_number:  # Assuming 'Zipline Order #' is in Column A
                row[1].value = order_numbers_str  # Assuming 'Cust Order #' is in Column B
                row[8].value = po_numbers_str  # Assuming 'PO and Confirmation Number' is Column I

        # Save the workbook to the Excel file
        self.workbook.save(filename=r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")

        # Start a timer to call show_next_form after a delay
        QTimer.singleShot(100, self.show_next_form)

    def open_pdf(self, invoice_number):
        # Define the base directory and the possible subdirectories
        base_dir = r"C:\Users\Andy Weigl\Kodiak Cakes\Kodiak Cakes Team Site - Public\Vendors\Zipline Logistics\Bills\2023"
        sub_dirs = [f"{i:02}.2023" for i in range(12, 0, -1)]  # 12.2023, 11.2023, ..., 01.2023

        # Try to find and open the PDF file
        for sub_dir in sub_dirs:
            full_dir = os.path.join(base_dir, sub_dir)
            if os.path.exists(full_dir):
                for file in os.listdir(full_dir):
                    # Check if the file is a PDF and its name contains the invoice number
                    if file.endswith(".pdf") and str(invoice_number) in file:
                        full_path = os.path.join(full_dir, file)
                        process = subprocess.Popen([r"C:\Program Files\Nitro\Pro\13\NitroPDF.exe", full_path])
                        return process

        # PDF file not found
        print(f"Could not find PDF file for invoice number {invoice_number}")
        return None

# Load the workbook and get the active worksheet
workbook = load_workbook(filename=r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")
worksheet = workbook.active

# Convert the worksheet to a list of dictionaries for easier processing
data = []
for row in worksheet.iter_rows(min_row=2, values_only=True):  # min_row=2 to skip the header
    data.append({
        'Zipline Order #': row[0],
        'Cust Order #': row[1],
        'PO and Confirmation Number': row[8],
    })

# Remove duplicate invoice numbers
seen = set()
data = [item for item in data if not (item['Zipline Order #'] in seen or seen.add(item['Zipline Order #']))]

# Initialize QApplication
app = QApplication([])

# Create the form manager and start the process
form_manager = FormManager(data)
form_manager.start()

app.exec_()




# Load the workbook
workbook = load_workbook(filename=r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")

# Select the "Data" worksheet
worksheet = workbook["Data"]

# Iterate over all rows in column I (9) and copy the values in column U (21)
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=9, max_col=9):
    for cell in row:
        # Copy value from column I to column U
        worksheet.cell(row=cell.row, column=21, value=cell.value)

# Iterate over all rows in column I (9) and copy the values in column U (21)
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=2, max_col=2):
    for cell in row:
        # Copy value from column I to column U
        worksheet.cell(row=cell.row, column=22, value=cell.value)

# Save the workbook
workbook.save(filename=r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")

# Step 6: Split out multiple PO numbers and Order numbers, and adjust the amounts for splits

# Load workbook and select the "Data" sheet
workbook = openpyxl.load_workbook(r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")
sheet = workbook["Data"]

# Create a new temporary sheet
temp_sheet = workbook.create_sheet("Temp")

# Copy headers
for idx, cell in enumerate(list(sheet.rows)[0], start=1):
    temp_sheet.cell(row=1, column=idx).value = cell.value

# Keep track of the new sheet row
new_row = 2

# For each row in the original sheet
for row in list(sheet.rows)[1:]:
    # Split order numbers and PO numbers by comma
    order_numbers = str(row[1].value).split(", ")
    po_numbers = str(row[8].value).split(", ")

    # Take the max length of the two lists
    max_len = max(len(order_numbers), len(po_numbers))

    # Extend the lists to the max length
    if len(order_numbers) != max_len:
        order_numbers.extend([order_numbers[0]] * (max_len - len(order_numbers)))
    if len(po_numbers) != max_len:
        po_numbers.extend([po_numbers[0]] * (max_len - len(po_numbers)))

    # Calculate the charge amount per order/PO number
    original_charge = float(row[10].value)
    charge_per_order = original_charge / max_len

    # Container for the actual charges after rounding
    actual_charges = []

    # For each order number and PO number
    for idx, (order_number, po_number) in enumerate(zip(order_numbers, po_numbers)):
        # Copy the entire row to the new sheet
        for col_idx, cell in enumerate(row, start=1):
            temp_sheet.cell(row=new_row, column=col_idx).value = cell.value

        # Replace the order number and PO number in the new sheet
        temp_sheet.cell(row=new_row, column=2).value = order_number
        temp_sheet.cell(row=new_row, column=9).value = po_number

        # Store the charge after rounding
        actual_charges.append(round(charge_per_order, 2))

        # Increment the new row counter
        new_row += 1

    # If there's a discrepancy due to rounding, adjust the final charge
    if sum(actual_charges) != original_charge:
        actual_charges[-1] += original_charge - sum(actual_charges)

    # Update the charges in the new sheet with the actual charges
    for i, actual_charge in enumerate(actual_charges, start=new_row-len(actual_charges)):
        temp_sheet.cell(row=i, column=11).value = round(actual_charge, 2)

# Delete the original sheet
workbook.remove(sheet)

# Rename the temporary sheet to the original sheet's name
temp_sheet.title = "Data"

# Move "Data" sheet to the first position
workbook._sheets.sort(key=lambda ws: ws.title!='Data')

# Save the workbook
workbook.save(r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")


# Step 7: filling in item, expense, line memo, and department

def set_location(ws):
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        # Check that cell values are not "None" before trying to access them
        col_C = str(row[2].value).lower() if row[2].value is not None else ""
        col_F = str(row[5].value).lower() if row[5].value is not None else ""
        col_G = str(row[6].value).lower() if row[6].value is not None else ""
        col_H = str(row[7].value).lower() if row[7].value is not None else ""

        # conditions
        if col_C == "us waffle co" and col_G == "carthage":
            row[12].value = "Carthage - Americold"
        elif col_C == "us waffle co" and col_G == "allentown":
            row[12].value = "Allentown - Americold"
        elif col_C == "us waffle co" and col_G == "atlanta":
            row[12].value = "Atlanta - Americold"
        elif col_C == "us waffle co" and col_G == "clearfield":
            row[12].value = "Clearfield - Americold"
        elif "element" in col_C and "bj" in col_F:
            row[12].value = "Element Food Solutions"
            row[13].value = "BJs Wholesale Club"
        elif "element" in col_C and "costco" in col_F and "nj" in col_H:
            row[12].value = "Element Food Solutions"
            row[13].value = "Costco Wholesale : Monroe"
        elif "element" in col_C and "costco" in col_F and "md" in col_H:
            row[12].value = "Element Food Solutions"
            row[13].value = "Costco Wholesale : Frederick"
        elif "element" in col_C and "costco" in col_F and "tx" in col_H:
            row[12].value = "Element Food Solutions"
            row[13].value = "Costco Wholesale : Katy"
        elif "element" in col_C and "costco" in col_F and "fl" in col_H:
            row[12].value = "Element Food Solutions"
            row[13].value = "Costco Wholesale : West Palm Beach"
        elif "element" in col_C and "costco" in col_F and "ga" in col_H:
            row[12].value = "Element Food Solutions"
            row[13].value = "Costco Wholesale : Atlanta"
        elif "honeyville" in col_C and "costco" in col_F and "fl" in col_H:
            row[12].value = "Honeyville Dropship"
            row[13].value = "Costco Wholesale : West Palm Beach"
        elif "honeyville" in col_C and "costco" in col_F and "nj" in col_H:
            row[12].value = "Honeyville Dropship"
            row[13].value = "Costco Wholesale : Monroe"
        elif "honeyville" in col_C and "costco" in col_F and "md" in col_H:
            row[12].value = "Honeyville Dropship"
            row[13].value = "Costco Wholesale : Frederick"
        elif "honeyville" in col_C and "costco" in col_F and "tx" in col_H:
            row[12].value = "Honeyville Dropship"
            row[13].value = "Costco Wholesale : Katy"
        elif "honeyville" in col_C and "costco" in col_F and "ga" in col_H:
            row[12].value = "Honeyville Dropship"
            row[13].value = "Costco Wholesale : Atlanta"
        elif "rjw" in col_F and "10" in col_F:
            row[12].value = "RJW Logistics W10"
        elif "rjw" in col_F and "12" in col_F:
            row[12].value = "RJW Logistics W12"
        elif "rjw" in col_F:
            row[12].value = "RJW Logistics W10"
        elif col_G == "laredo":
            row[12].value = "Cerealto - Laredo"
        elif "schulze" in col_F:
            row[12].value = "Schulze & Burch Biscuit Co."
        elif "honeyville" in col_F:
            row[12].value = "Honeyville Dropship"
        elif "magi" in col_F:
            row[12].value = "Magi Foods"
        elif "hearthside des plaines p4 plant" in col_F:
            row[12].value = "Hearthside Food Solutions - Des Plaines"
        elif "element food solutions" in col_F:
            row[12].value = "Element Food Solutions"
        elif "honeyville" in col_C:
            row[12].value = "Honeyville Dropship"
        elif "safeway" in col_C:
            row[12].value = "RJW Logistics W10"


def set_partial_memo(ws):
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        # Python columns start from 0, so column C is 2, F is 5, G is 6, and M is 12

        # Check that cell values are not "None" before trying to access them
        col_B = str(row[1].value).lower() if row[1].value is not None else ""
        col_I = str(row[8].value).lower() if row[8].value is not None else ""
        col_J = str(row[9].value).lower() if row[9].value is not None else ""
        col_C = str(row[2].value).lower() if row[2].value is not None else ""
        col_F = str(row[5].value).lower() if row[5].value is not None else ""
        col_G = str(row[6].value).lower() if row[6].value is not None else ""
        col_H = str(row[7].value).lower() if row[7].value is not None else ""
        col_M = str(row[12].value).lower() if row[12].value is not None else ""
        col_N = str(row[13].value).lower() if row[13].value is not None else ""
        col_O = str(row[14].value).lower() if row[14].value is not None else ""
        col_P = str(row[15].value).lower() if row[15].value is not None else ""
        col_Q = str(row[16].value).lower() if row[16].value is not None else ""
        col_R = str(row[17].value).lower() if row[17].value is not None else ""
        col_S = str(row[18].value).lower() if row[18].value is not None else ""
        col_T = str(row[19].value).lower() if row[19].value is not None else ""

        if "americold" in col_M and "carthage" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " " + str(row[12].value)
            row[17].value = "Landed Cost - Americold - Carthage - Freight"
            row[19].value = "Operations : Supply Chain"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " " + str(row[12].value)
        elif "americold" in col_M and "atlanta" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " " + str(row[12].value)
            row[17].value = "Landed Cost - Americold - Atlanta - Freight"
            row[19].value = "Operations : Supply Chain"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " " + str(row[12].value)
        elif "americold" in col_M and "allentown" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " " + str(row[12].value)
            row[17].value = "Landed Cost - Americold - Allentown - Freight"
            row[19].value = "Operations : Supply Chain"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " " + str(row[12].value)
        elif "americold" in col_M and "clearfield" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " " + str(row[12].value)
            row[17].value = "Landed Cost - Americold - Clearfield - Freight"
            row[19].value = "Operations : Supply Chain"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " " + str(row[12].value)
        elif "rjw" in col_M and "10" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " RJWW10"
            row[17].value = "Landed Cost - RJW - Freight"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " RJWW10"
        elif "rjw" in col_M and "12" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " RJWW12"
            row[17].value = "Landed Cost - RJW - Freight"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " RJWW12"
        elif col_M == "element food solutions" and col_N == "bjs wholesale club":
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[13].value)
            row[18].value = "62006 Shipping Expense"
            row[19].value = "Sales"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[13].value)
        elif col_M == "element food solutions" and "costco" in col_N:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[13].value)
            row[18].value = "62006 Shipping Expense"
            row[19].value = "Sales"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[13].value)
        elif col_M == "honeyville dropship" and "honeyville" in col_F:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif col_C == "msi express inc." and col_F == "element food solutions":
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif col_F == "magi foods" and "massey" in col_C:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif col_F == "magi foods" and "mac" in col_C:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif col_F == "magi foods" and "batory" in col_C:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif col_F == "element food solutions" and "cerealto" in col_C:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif col_F == "element food solutions" and "cerealto" in col_C:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif col_M == "honeyville dropship" and "costco" in col_N:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[13].value)
            row[18].value = "62006 Shipping Expense"
            row[19].value = "Sales"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[13].value)
        elif "cerealto" in col_F and "bay state" in col_C:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif col_C == "visstun cups":
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif "schulze" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif "magi foods" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif "hearthside" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif "element" in col_C and "element" in col_F:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif "cerealto" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"
        elif "honeyville" in col_M:
            row[14].value = "PO" + str(row[8].value) + "; " + str(row[1].value) + " Ingredient Transfer"
            row[18].value = "50004 Cost of Goods Sold : Manual"
            row[19].value = "Operations : Manufacturing"
            row[22].value = "PO" + str(row[20].value) + "; " + str(row[21].value) + " Ingredient Transfer"


def set_final_memo(ws):
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        # Check that cell values are not "None" before trying to access them
        col_J = str(row[9].value).lower() if row[9].value is not None else ""
        col_M = str(row[12].value).lower() if row[12].value is not None else ""

        if col_J != "freight charge":
            row[15].value = str(row[14].value) + " - " + str(row[9].value)
        elif col_J == "freight charge":
            row[15].value = str(row[14].value)

        if col_M == "carthage - americold":
            row[16].value = "Retail : Frozen"
        elif col_M == "allentown - americold":
            row[16].value = "Retail : Frozen"
        elif col_M == "clearfield - americold":
            row[16].value = "Retail : Frozen"
        elif col_M == "atlanta - americold":
            row[16].value = "Retail : Frozen"

def main1():
    wb_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx"
    wb = openpyxl.load_workbook(wb_path)
    ws = wb['Data']

    # Change the header of various columns
    ws["M1"] = "Location"
    ws["N1"] = "Customer"
    ws["O1"] = "Partial Line Memo"
    ws["P1"] = "Line Memo with Charges"
    ws["Q1"] = "Division"
    ws["R1"] = "Item"
    ws["S1"] = "Expense"
    ws["T1"] = "Department"
    ws["W1"] = "Primary PO"

    # hide column O
    ws.column_dimensions['O'].hidden = True

    set_location(ws)
    set_partial_memo(ws)
    set_final_memo(ws)

    wb.save(wb_path)


if __name__ == "__main__":
    main1()


class DataHandler:
    # Handles reading and writing data from/to Excel
    def __init__(self, file_name):
        self.file_name = file_name
        self.df = self._read_file()

    def _read_file(self):
        try:
            return pd.read_excel(self.file_name, engine="openpyxl")
        except Exception as e:
            print(f"Failed to read file: {e}")
            return None

    def get_unique_po_numbers(self):
        if self.df is not None:
            # Filter out rows where "Divisions" column is not empty
            df = self.df[self.df["Division"].isna()]
            df = df.drop_duplicates(subset=["PO and Confirmation Number", "Zipline Order #"])
            po_numbers = df["PO and Confirmation Number"].tolist()
            invoice_numbers = df["Zipline Order #"].tolist()
            po_invoice_dict = {str(po): invoice for po, invoice in zip(po_numbers, invoice_numbers)}
            return po_invoice_dict
        else:
            return {}

    def update_file_with_divisions(self, divisions):
        try:
            # Convert keys in divisions to strings
            divisions = {str(key): value for key, value in divisions.items()}
            book = load_workbook(filename=r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")
            sheet = book["Data"]
            for index, row in self.df.iterrows():
                po = str(row["PO and Confirmation Number"])
                if po in divisions:
                    cell = sheet.cell(row=index + 2, column=17)
                    cell.value = divisions[po]
            book.save(filename=r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")
        except Exception as e:
            # Create a root Tkinter window and hide it
            root = tk.Tk()
            root.withdraw()
            # Show the error messagebox
            messagebox.showerror("Error", f"Failed to update file: {e}")
            # Destroy the root window
            root.destroy()


class WindowHandler:
    # Handles interactions with the window
    @staticmethod
    def find_and_activate_window(title_substr):
        for window in gw.getAllWindows():
            if title_substr in window.title:
                window.activate()
                return True
        return False

    @staticmethod
    def open_url(url):
        webbrowser.open(url)
        while not WindowHandler.find_and_activate_window("NetSuite (Kodiak Cakes LLC) — Mozilla Firefox"):
            pass


class GUIHandler:
    # Handles GUI creation and updates
    def __init__(self, po_invoice_dict, divisions_list):
        self.root = tk.Tk()
        self.root.title("Divisions")
        self.root.attributes('-topmost', True)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.current_index = 0
        self.po_invoice_dict = po_invoice_dict
        self.po_numbers = list(self.po_invoice_dict.keys())  # Extract the PO numbers from the dict
        self.current_po = tk.StringVar()
        self.current_po.set(self.po_numbers[self.current_index] if self.po_numbers else "")
        self.divisions = {}
        self.divisions_list = divisions_list  # Save divisions_list as an instance variable
        self.stop_checking_clipboard = False
        self.clipboard_check_callback_id = None
        self.gui_destroyed = False
        self.pdf_process = None  # Store the subprocess for the PDF viewer here

        # Create a custom style for the combobox
        self.custom_style = ttk.Style()
        self.custom_style.map('Flashing.TCombobox',
                              fieldbackground=[('readonly', 'white')],
                              selectbackground=[('readonly', 'white')]
                              )

        # GUI elements
        entry_width = 23
        self.po_label = ttk.Label(self.root, text="PO Number:")
        self.po_label.grid(column=0, row=0, pady=5, padx=5, sticky='e')

        self.po_entry = ttk.Entry(self.root, state="readonly", textvariable=self.current_po, width=entry_width)
        self.po_entry.grid(column=1, row=0, pady=5, padx=5)

        self.invoice_label = ttk.Label(self.root, text="Invoice Number:")
        self.invoice_label.grid(column=0, row=1, pady=5, padx=5, sticky='e')

        self.current_invoice = tk.StringVar()  # Add a StringVar for the invoice number
        self.invoice_entry = ttk.Entry(self.root, state="readonly", textvariable=self.current_invoice, width=entry_width)
        self.invoice_entry.grid(column=1, row=1, pady=5, padx=5)

        button_width = 15
        self.copy_button = ttk.Button(self.root, text="Copy", command=self.copy_po, width=button_width)
        self.copy_button.grid(column=2, row=0, pady=5, padx=5)

        self.pdf_button = ttk.Button(self.root, text="View/Hide PDF", command=self.view_hide_pdf, width=button_width)
        self.pdf_button.grid(column=2, row=1, pady=5, padx=5)

        self.division_label = ttk.Label(self.root, text="Division:")
        self.division_label.grid(column=0, row=2, pady=5, padx=5, sticky='e')

        self.division_options = [
            "Club", "Club : Baking", "Club : Bars", "Club : Crackers", "Club : Cups", "Club : Frozen", "Club : Oatmeal",
            "Club : Other", "Club : Pancake", "Club : Syrup", "Retail : Baking", "Retail : Bars", "Retail : Cookies",
            "Retail : Crackers", "Retail : Cups", "Retail : Frozen", "Retail : Oatmeal", "Retail : Other",
            "Retail : Pancake", "Retail : Syrup"
        ]

        self.division_var = tk.StringVar()  # variable to hold selected option

        self.division_combo = ttk.Combobox(self.root, textvariable=self.division_var, style='Flashing.TCombobox')
        self.division_combo['values'] = self.division_options
        self.division_combo.bind('<KeyRelease>', self.check_key)
        self.division_combo.bind('<Return>', lambda event: self.next_po())
        self.division_combo.grid(column=1, row=2, pady=5, padx=5)

        self.next_button = ttk.Button(self.root, text="Next", command=self.next_po, width=button_width)
        self.next_button.grid(column=2, row=2, pady=5, padx=5)  # Moved to row 2

        style = ttk.Style()
        style.configure("TButton", font=("Arial", 10))
        style.map("TButton", foreground=[("alternate", "black"), ("!alternate", "black")])

        self.status_label = tk.Label(self.root, text="", font=("Arial", 10))
        self.status_label.grid(row=3, column=0, pady=5, columnspan=2)  # Moved to row 3



    def check_key(self, event=None):
        if event.keysym != 'Return':  # Check if the key is not "Enter"
            value = self.division_var.get().lower()
            if value == '':
                data = self.division_options
            else:
                data = []
                for item in self.division_options:
                    if value in item.lower():
                        data.append(item)

            self.division_combo['values'] = data
            self.division_combo.event_generate('<Down>')  # Show the dropdown

    def view_hide_pdf(self):
        if self.pdf_process is None:
            # Open the PDF and store the subprocess
            invoice_number = self.po_invoice_dict[self.current_po.get()]
            self.pdf_process = self.open_pdf(invoice_number)

            # Check if the PDF opening was successful
            if self.pdf_process is None:
                print("Error: Failed to open PDF.")
        else:
            # Kill the subprocess to "hide" the PDF and reset self.pdf_process
            if self.pdf_process is not None:
                self.pdf_process.kill()
                self.pdf_process = None
            else:
                print("Error: No PDF process to kill.")

    @staticmethod
    def open_pdf(invoice_number):
        # Define the base directory and the possible subdirectories
        base_dir = r"C:\Users\Andy Weigl\Kodiak Cakes\Kodiak Cakes Team Site - Public\Vendors\Zipline Logistics\Bills\2023"
        sub_dirs = [f"{i:02}.2023" for i in range(12, 0, -1)]  # 12.2023, 11.2023, ..., 01.2023

        # Try to find and open the PDF file
        for sub_dir in sub_dirs:
            full_dir = os.path.join(base_dir, sub_dir)
            if os.path.exists(full_dir):
                for file in os.listdir(full_dir):
                    # Check if the file is a PDF and its name contains the invoice number
                    if file.endswith(".pdf") and str(invoice_number) in file:
                        full_path = os.path.join(full_dir, file)
                        process = subprocess.Popen([r"C:\Program Files\Nitro\Pro\13\NitroPDF.exe", full_path])
                        return process

        # PDF file not found
        return None

    def run(self):
        self.update_po()
        self.check_clipboard()
        self.root.mainloop()

    def copy_po(self):
        po_value = "PO" + self.current_po.get()
        pyperclip.copy(po_value)
        self.copy_button.config(text="Copied!")

    def next_po(self):
        division = self.division_combo.get().rstrip()
        self.divisions[self.current_po.get()] = division

        if self.current_index < len(self.po_numbers) - 1:
            self.current_index += 1
            self.update_po()
            self.copy_button.config(text="✓")
            self.root.after(1500, self.reset_copy_button_text)
            self.division_combo.set('')  # Clear the division combo box
            if self.pdf_process is not None:
                # If we're moving to the next PO, close the PDF for the current PO
                self.pdf_process.kill()
                self.pdf_process = None
        else:
            self.stop_checking_clipboard = True
            self.root.destroy()

    def reset_copy_button_text(self):
        if pyperclip.paste() == self.current_po.get():
            self.copy_button.config(text="Copied!", style="TButton")
            self.copy_button.state(["alternate"])
        else:
            self.copy_button.config(text="Copy", style="TButton")
            self.copy_button.state(["!alternate"])

    def update_po(self):
        if self.current_index < len(self.po_numbers):
            current_po = self.po_numbers[self.current_index]
            self.current_po.set(current_po)  # Set the current PO number
            self.current_invoice.set(self.po_invoice_dict[current_po])  # Set the current invoice number
            self.copy_po()  # Copy current PO number to the clipboard
            self.division_var.set('')  # Clear the division combo box

    def on_closing(self):
        self.stop_checking_clipboard = True
        self.gui_destroyed = True
        self.root.destroy()
        if self.pdf_process is not None:
            self.pdf_process.kill()
            self.pdf_process = None
        else:
            print("Error: No PDF process to kill.")

    def check_clipboard(self):
        if self.stop_checking_clipboard:
            return

        clipboard_content = pyperclip.paste()
        if clipboard_content != self.current_po.get():
            self.copy_button.config(text="Copy")

        if clipboard_content in self.divisions_list:
            self.division_combo.set(clipboard_content)  # Use division_combo.set() to set the division combo box text
            accepted_po = self.current_po.get()
            if self.current_index < len(self.po_numbers) - 1:
                self.next_po()
            elif self.current_index == len(self.po_numbers) - 1:  # If we are at the final PO number
                self.stop_checking_clipboard = True  # Stop checking the clipboard
                self.next_po()  # Proceed to the next PO number

            if self.status_label.winfo_exists() and not self.gui_destroyed:
                self.status_label.config(text=f"                      Division accepted - {accepted_po}", fg="green")
            self.flash_entry_background()

        # Only schedule the function to run again if we are not at the final PO number
        if self.root is not None and self.root.winfo_exists() and not self.stop_checking_clipboard:
            self.clipboard_check_callback_id = self.root.after(250, self.check_clipboard)

    def flash_entry_background(self):
        # Change the style to create a flashing effect
        self.custom_style.map('Flashing.TCombobox',
                              fieldbackground=[('readonly', 'green')],
                              selectbackground=[('readonly', 'green')]
                              )
        self.root.after(1500, lambda: self.custom_style.map('Flashing.TCombobox',
                                                            fieldbackground=[('readonly', 'white')],
                                                            selectbackground=[('readonly', 'white')]
                                                            )
                        )



def main2():
    file_name = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx"

    data_handler = DataHandler(file_name)
    po_invoice_dict = data_handler.get_unique_po_numbers()
    window_found = WindowHandler.find_and_activate_window("NetSuite (Kodiak Cakes LLC) — Mozilla Firefox")
    if not window_found:
        WindowHandler.open_url("https://4289607.app.netsuite.com/app/login/secure/enterpriselogin.nl?c=4289607&whence=")
    gui_handler = GUIHandler(po_invoice_dict, divisions_list)
    gui_handler.run()
    data_handler.update_file_with_divisions(gui_handler.divisions)
    show_message_box("Success", "All divisions have been entered!")
    return  # This will end the execution of the function


divisions_list = [
    "Category Report ",
    "Club ",
    "Club : Baking ",
    "Club : Bars ",
    "Club : Crackers ",
    "Club : Cups ",
    "Club : Frozen ",
    "Club : Oatmeal ",
    "Club : Other ",
    "Club : Pancake ",
    "Club : Syrup ",
    "E-commerce ",
    "E-commerce : Baking ",
    "E-commerce : Bars ",
    "E-commerce : Cookies ",
    "E-commerce : Crackers ",
    "E-commerce : Cups ",
    "E-commerce : Oatmeal ",
    "E-commerce : Pancake ",
    "E-commerce : Syrup ",
    "Retail ",
    "Retail : Baking ",
    "Retail : Bars ",
    "Retail : Cookies ",
    "Retail : Crackers ",
    "Retail : Cups ",
    "Retail : Frozen ",
    "Retail : Oatmeal ",
    "Retail : Other ",
    "Retail : Pancake ",
    "Retail : Syrup ",
]


main2()


# Load the workbook
wb = load_workbook(r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")

# Get the source and destination worksheets
ws_data = wb["Data"]
ws_upload = wb["Zipline Upload Template"]


# Clear all rows below the header in the destination worksheet
for row in ws_upload['A2:AA' + str(ws_upload.max_row)]:
    for cell in row:
        cell.value = None

location_dictionary = {
    'Allentown - Americold': 1,
    'Amazon': 22,
    'AST': 51,
    'Atlanta - Americold': 21,
    'Carthage - Americold': 20,
    'Cerealto - Hidalgo': 50,
    'Cerealto - Laredo': 48,
    'CII Foods': 36,
    'Clearfield - Americold': 15,
    'Element Food Solutions': 14,
    'Fantasy Cookie Corporation': 30,
    'Fortis Solutions Group': 52,
    'Hearthside Dropship': 28,
    'Hearthside Food Solutions': 16,
    'Hearthside Food Solutions - Des Plaines': 37,
    'Honeyville Dropship': 9,
    'Magi Foods': 29,
    'MSI Express': 34,
    'Park City HQ': 10,
    'Probst Farms': 2,
    'Red Stag Fulfillment': 27,
    'Redeployment In-Transit': 23,
    'RJW Logistics W10': 31,
    'RJW Logistics W12': 47,
    'Schulze & Burch Biscuit Co.': 49,
    'Treehouse': 24,
    'Treehouse - Brantford': 32,
    'Treehouse - Medina': 35,
    'US Waffle': 6,
}

customer_dictionary = {
    'BJs Wholesale Club': 422,
    'Costco Wholesale': 328,
    'Costco Wholesale : Atlanta': 1963,
    'Costco Wholesale : Aurora': 2335,
    'Costco Wholesale : Dallas': 1964,
    'Costco Wholesale : Frederick': 1965,
    'Costco Wholesale : Katy': 1966,
    'Costco Wholesale : Laredo': 1967,
    'Costco Wholesale : Mira Loma': 1968,
    'Costco Wholesale : Monroe': 1969,
    'Costco Wholesale : Morris': 1970,
    'Costco Wholesale : Owatonna': 2267,
    'Costco Wholesale : Roadshow': 1976,
    'Costco Wholesale : Salt Lake': 1971,
    'Costco Wholesale : Sumner': 1972,
    'Costco Wholesale : Tolleson': 1973,
    'Costco Wholesale : Tracy': 1974,
    'Costco Wholesale : Van Buren': 1975,
    'Costco Wholesale : West Palm Beach': 1993,
    'Costco Wholesale Canada': 428,
}

expense_dictionary = {
    '50004 Cost of Goods Sold : Manual': 426,
    '62006 Shipping Expense': 294,
}

item_dictionary = {
    'Landed Cost - Americold - Allentown - Freight': 1481,
    'Landed Cost - Americold - Allentown - Handling': 1482,
    'Landed Cost - Americold - Atlanta - Freight': 1512,
    'Landed Cost - Americold - Atlanta - Handling': 1513,
    'Landed Cost - Americold - Carthage - Freight': 1510,
    'Landed Cost - Americold - Carthage - Handling': 1511,
    'Landed Cost - Americold - Clearfield - Freight': 1489,
    'Landed Cost - Americold - Clearfield - Handling': 1490,
    'Landed Cost - Americold - Dallas - Freight': 1487,
    'Landed Cost - Americold - Dallas - Handling': 1488,
    'Landed Cost - Americold - Salt Lake - Freight': 2640,
    'Landed Cost - Americold - Salt Lake - Handling': 2641,
    'Landed Cost - AST - Freight': 1479,
    'Landed Cost - AST - Handling': 1480,
    'Landed Cost - IWI - Freight': 1485,
    'Landed Cost - IWI - Handling': 1486,
    'Landed Cost - RJW - Freight': 1483,
    'Landed Cost - RJW - Handling': 1484,
}

# Start copying from the second row in column A to exclude the header
for i in range(2, ws_data.max_row + 1):  # Iterate from the second row until the last row of data
    # If the cell in column A is not empty
    if ws_data.cell(row=i, column=1).value:
        # 1=A  2=B  3=C  4=D  5=E  6=F  7=G  8=H  9=I  10=J  11=K  12=L  13=M  14=N  15=O  16=P
        # 17=Q  18=R  19=S  20=T  21=U  22=V  23=W  24=X  25=Y  26=Z
        ws_upload.cell(row=i, column=5).value = ws_data.cell(row=i, column=14).value  # Customer
        ws_upload.cell(row=i, column=7).value = ws_data.cell(row=i, column=17).value  # Division
        ws_upload.cell(row=i, column=8).value = ws_data.cell(row=i, column=20).value  # Department
        ws_upload.cell(row=i, column=10).value = ws_data.cell(row=i, column=12).value  # Date
        ws_upload.cell(row=i, column=11).value = ws_data.cell(row=i, column=19).value  # Expenses
        ws_upload.cell(row=i, column=13).value = ws_data.cell(row=i, column=18).value  # Item
        ws_upload.cell(row=i, column=2).value = ws_data.cell(row=i, column=13).value  # Location
        ws_upload.cell(row=i, column=4).value = ws_data.cell(row=i, column=9).value  # PO
        ws_upload.cell(row=i, column=15).value = ws_data.cell(row=i, column=11).value  # Amount
        ws_upload.cell(row=i, column=16).value = ws_data.cell(row=i, column=16).value  # Line Memo
        ws_upload.cell(row=i, column=9).value = ws_data.cell(row=i, column=1).value  # Reference Number
        ws_upload.cell(row=i, column=1).value = "1940"
        name = ws_upload.cell(row=i, column=2).value
        if name in location_dictionary:
            ws_upload.cell(row=i, column=3).value = location_dictionary[name]  # Location Internal ID
        name = ws_upload.cell(row=i, column=5).value
        if name in customer_dictionary:
            ws_upload.cell(row=i, column=6).value = customer_dictionary[name]  # Customer Internal ID
        name = ws_upload.cell(row=i, column=11).value
        if name in expense_dictionary:
            ws_upload.cell(row=i, column=12).value = expense_dictionary[name]  # Expense Internal ID
        name = ws_upload.cell(row=i, column=13).value
        if name in item_dictionary:
            ws_upload.cell(row=i, column=14).value = item_dictionary[name]  # Item Internal ID

# Save the workbook
wb.save(r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")

excel_file_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx"

os.startfile(excel_file_path)

# Wait for 2 seconds
time.sleep(2)

show_message_box("Review", "Please open the excel file and review. Then, make sure the workbook is saved and press ` to continue...")

# Wait for the Enter key to be pressed globally
keyboard.wait('`')

def clear_sheet(sheet):
    sheet.delete_rows(1, sheet.max_row)
    sheet.delete_cols(1, sheet.max_column)


def copy_column(source_sheet, source_column, target_sheet, target_column):
    for row in range(1, source_sheet.max_row + 1):
        target_sheet.cell(row=row, column=target_column).value = source_sheet.cell(row=row, column=source_column).value


def remove_duplicates(sheet, col_index, row_offset=1):
    unique_values = set()
    row_numbers_to_delete = []

    for row in sheet.iter_rows(min_row=row_offset + 1):
        cell_value = row[col_index].value
        if cell_value not in unique_values:
            unique_values.add(cell_value)
        else:
            row_numbers_to_delete.append(row[col_index].row)

    for row_number in reversed(row_numbers_to_delete):
        sheet.delete_rows(row_number)

    return sheet


workbook = openpyxl.load_workbook(r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")

# Clear the contents of the "Hub Group Primary" sheet
zipline_primary_sheet_name = 'Zipline Primary'
if zipline_primary_sheet_name in workbook.sheetnames:
    zipline_primary_sheet = workbook[zipline_primary_sheet_name]
    clear_sheet(zipline_primary_sheet)

# Clear the contents of the "Hub Group Expenses" sheet
zipline_expenses_sheet_name = 'Zipline Expenses'
if zipline_expenses_sheet_name in workbook.sheetnames:
    zipline_expenses_sheet = workbook[zipline_expenses_sheet_name]
    clear_sheet(zipline_expenses_sheet)

# Clear the contents of the "Hub Group Items" sheet
zipline_items_sheet_name = 'Zipline Items'
if zipline_items_sheet_name in workbook.sheetnames:
    zipline_items_sheet = workbook[zipline_items_sheet_name]
    clear_sheet(zipline_items_sheet)

# Get the "Upload Template" worksheet
upload_template_sheet_name = 'Zipline Upload Template'
upload_template_sheet = workbook[upload_template_sheet_name]
data_template_sheet_name = "Data"
data_template_sheet = workbook[data_template_sheet_name]

# Copy the specified columns from the "Upload Template" worksheet to the "Hub Group Primary" worksheet
copy_column(upload_template_sheet, 1, zipline_primary_sheet, 1)  # Copy Column A to Column A
copy_column(upload_template_sheet, 10, zipline_primary_sheet, 2)  # Copy Column J to Column B
copy_column(upload_template_sheet, 9, zipline_primary_sheet, 3)  # Copy Column I to Column C
copy_column(data_template_sheet, 23, zipline_primary_sheet, 4)  # Copy Column P to Column D

# Remove duplicates based on Column C in the "Hub Group Primary" worksheet
remove_duplicates(zipline_primary_sheet, 2, 1)

# Notify the user to select the email with the Excel file
print("Wait, then press ` to continue...")

# Wait for the Enter key to be pressed globally
keyboard.wait('`')

# Get the source and destination worksheets
upload_sheet = workbook["Zipline Upload Template"]
expenses_sheet = workbook["Zipline Expenses"]
items_sheet = workbook["Zipline Items"]

# Copy headers from the upload sheet to the expenses sheet
for col_num in range(1, upload_sheet.max_column + 1):
    expenses_sheet.cell(row=1, column=col_num).value = upload_sheet.cell(row=1, column=col_num).value

# Find the first empty row in the expenses sheet
first_empty_row_expenses = expenses_sheet.max_row + 1 if expenses_sheet.max_row > 1 else 2

# Copy rows from the upload sheet to the expenses sheet
for i in range(2, upload_sheet.max_row + 1):  # Start from the second row to exclude the header
    # If the cell in column A (1) is not empty
    if upload_sheet.cell(row=i, column=1).value:
        # If the cell in column K (11) is not empty
        if upload_sheet.cell(row=i, column=11).value:
            # Copy the entire row to the expenses sheet
            for j in range(1, upload_sheet.max_column + 1):
                expenses_sheet.cell(row=first_empty_row_expenses, column=j).value = upload_sheet.cell(row=i, column=j).value
            # Move to the next row in the expenses sheet
            first_empty_row_expenses += 1

# Copy headers from the upload sheet to the items sheet
for col_num in range(1, upload_sheet.max_column + 1):
    items_sheet.cell(row=1, column=col_num).value = upload_sheet.cell(row=1, column=col_num).value

# Find the first empty row in the items sheet
first_empty_row_items = items_sheet.max_row + 1 if items_sheet.max_row > 1 else 2

# Copy rows from the upload sheet to the items sheet
for i in range(2, upload_sheet.max_row + 1):  # Start from the second row to exclude the header
    # If the cell in column A (1) is not empty
    if upload_sheet.cell(row=i, column=1).value:
        # If the cell in column K (11) is not empty
        if upload_sheet.cell(row=i, column=13).value:
            # Copy the entire row to the items sheet
            for j in range(1, upload_sheet.max_column + 1):
                items_sheet.cell(row=first_empty_row_items, column=j).value = upload_sheet.cell(row=i, column=j).value
            # Move to the next row in the items sheet
            first_empty_row_items += 1

workbook.save(r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx")

# Load the source workbook
source_workbook_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\Zipline Upload.xlsx"
source_workbook = load_workbook(source_workbook_path)

# Create date-based folder and file name
current_date = datetime.now().strftime("%m.%d")
csv_upload_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Zipline\CSV Upload"
date_folder_path = os.path.join(csv_upload_path, current_date)

# Create the directory if it doesn't already exist
os.makedirs(date_folder_path, exist_ok=True)

# Iterate over the list of worksheet names
for sheet_name in ["Zipline Primary", "Zipline Expenses", "Zipline Items"]:
    # Load the worksheet
    worksheet = source_workbook[sheet_name]

    # Set the output CSV file name using the worksheet name and current date
    output_csv = f"{sheet_name} - {current_date}.csv"

    # Set the full path of the output CSV file
    output_csv_path = os.path.join(date_folder_path, output_csv)

    # Save the entire worksheet to the output CSV file (overwriting it if it already exists)
    with open(output_csv_path, mode="w", newline="") as csv_file:
        csv_writer = csv.writer(csv_file)

        # Iterate through each row in the worksheet
        for row in worksheet.iter_rows():
            # Extract the values of the cells in the current row
            row_values = [cell.value for cell in row]
            # Write the row values to the CSV file
            csv_writer.writerow(row_values)

show_message_box("Success!", "The CSV files have been generated and are ready to be uploaded!")